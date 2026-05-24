"""
Schedule Extractor Service for KnowStrath RAG System

This service handles intelligent extraction of schedule data from Excel, PDF, and other
table-based documents, maintaining context awareness for class groups and time relationships.
"""

import re
import json
import logging
import hashlib
import pandas as pd
import openpyxl
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple
from dataclasses import dataclass
from datetime import datetime
import tabula
import openai
import os

logger = logging.getLogger("schedule_extractor")


@dataclass
class ScheduleEntry:
    """Represents a single schedule entry with all relevant information."""

    class_group: str
    subject: str
    unit_code: Optional[str]
    room: str
    day: str
    start_time: str
    end_time: str
    instructor: str
    session_type: str  # lecture, lab, practical
    semester: str
    raw_content: str  # Original cell content for debugging
    confidence: float = 1.0

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for storage."""
        return {
            "class_group": self.class_group,
            "subject": self.subject,
            "unit_code": self.unit_code,
            "room": self.room,
            "day": self.day,
            "start_time": self.start_time,
            "end_time": self.end_time,
            "instructor": self.instructor,
            "session_type": self.session_type,
            "semester": self.semester,
            "raw_content": self.raw_content,
            "confidence": self.confidence,
            "extraction_timestamp": datetime.now().isoformat(),
        }


class ScheduleExtractor:
    """Intelligent schedule extractor with table awareness and context propagation."""

    def __init__(self):
        """Initialize the schedule extractor."""
        self.openai_client = openai.OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

        # Schedule detection patterns
        self.schedule_indicators = [
            r"[A-Z]{2,6}\s+\d[A-Z].*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec).*20\d{2}",  # Class groups with semester
            r"\d{1,2}[:\.]\d{2}[-–]\d{1,2}[:\.]\d{2}",  # Time ranges
            r"(?:Monday|Tuesday|Wednesday|Thursday|Friday)",  # Days
            r"(?:LT|Lab|Room|MSB|SLS|STMB|Lecture Theatre)",  # Room patterns
            r"(?:ICS|HED|MATH|STAT)\s*\d{4}",  # Unit codes
            r"Dr\s+[A-Z][a-z]+",  # Instructor patterns
        ]

        # Time slot patterns
        self.time_patterns = [
            r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})",
            r"(\d{1,2})[:.](\d{2})\s*-\s*(\d{1,2})[:.](\d{2})",
        ]

        # Room/location patterns
        self.room_patterns = [
            r"LT\s*\d+",
            r"Lecture\s+Theatre\s+\d+",
            r"Lab\s*\d+",
            r"MSB\s+\d+",
            r"SLS\s+\w+",
            r"STMB\s+[\w\d\s]+",
            r"Room\s+\d+",
            r"RM\s+\d+",
        ]

    def detect_schedule_document(
        self, file_path: Union[str, Path], content: str = None
    ) -> Tuple[bool, float]:
        """
        Detect if a document contains schedule data.

        Args:
            file_path: Path to the document
            content: Optional text content to analyze

        Returns:
            Tuple of (is_schedule, confidence_score)
        """
        file_path = Path(file_path)

        # Check file type first
        if file_path.suffix.lower() in [".xlsx", ".xls"]:
            # Excel files with schedule-like names are likely schedules
            filename_lower = file_path.name.lower()
            if any(
                keyword in filename_lower
                for keyword in ["timetable", "schedule", "class", "ics"]
            ):
                return True, 0.9

        # Analyze content if provided
        if content:
            score = self._calculate_schedule_score(content)
            return score > 5, score / 10.0

        # Try to extract some content for analysis
        try:
            if file_path.suffix.lower() in [".xlsx", ".xls"]:
                df = pd.read_excel(file_path, nrows=20)  # Read first few rows
                content = df.to_string()
            elif file_path.suffix.lower() == ".pdf":
                # Try to extract tables from PDF
                tables = tabula.read_pdf(
                    file_path, pages=1, multiple_tables=True
                )
                if tables:
                    content = tables[0].to_string()

            if content:
                score = self._calculate_schedule_score(content)
                return score > 5, score / 10.0

        except Exception as e:
            logger.warning(
                f"Could not analyze content for schedule detection: {e}"
            )

        return False, 0.0

    def _calculate_schedule_score(self, content: str) -> int:
        """Calculate a score indicating how likely content is a schedule."""
        score = 0
        for pattern in self.schedule_indicators:
            matches = len(re.findall(pattern, content, re.IGNORECASE))
            if matches > 0:
                score += min(matches, 3)  # Cap contribution per pattern

        return score

    def extract_from_file(
        self, file_path: Union[str, Path]
    ) -> List[ScheduleEntry]:
        """
        Extract schedule entries from a file.

        Args:
            file_path: Path to the schedule file

        Returns:
            List of extracted schedule entries
        """
        file_path = Path(file_path)

        print(f"📊 Extracting schedule from: {file_path.name}")

        try:
            if file_path.suffix.lower() in [".xlsx", ".xls"]:
                return self._extract_from_excel(file_path)
            elif file_path.suffix.lower() == ".pdf":
                return self._extract_from_pdf(file_path)
            else:
                logger.warning(
                    f"Unsupported file type for schedule extraction: {file_path.suffix}"
                )
                return []

        except Exception as e:
            logger.error(f"Failed to extract schedule from {file_path}: {e}")
            return []

    def _extract_from_excel(self, file_path: Path) -> List[ScheduleEntry]:
        """Extract schedule from Excel file using openpyxl for merged-cell awareness."""
        schedules = []

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in wb.sheetnames:
                print(f"  📋 Processing sheet: {sheet_name}")
                ws = wb[sheet_name]
                rows = self._resolve_merged_cells(ws)
                sheet_schedules = self._extract_from_rows(rows, sheet_name)
                schedules.extend(sheet_schedules)

        except Exception as e:
            logger.error(f"openpyxl read failed ({e}), falling back to pandas")
            try:
                excel_file = pd.ExcelFile(file_path)
                for sheet_name in excel_file.sheet_names:
                    print(f"  📋 Processing sheet (pandas): {sheet_name}")
                    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                    rows = [list(row) for _, row in df.iterrows()]
                    sheet_schedules = self._extract_from_rows(rows, sheet_name)
                    schedules.extend(sheet_schedules)
            except Exception as e2:
                logger.error(f"Pandas fallback also failed: {e2}")

        print(f"✅ Extracted {len(schedules)} schedule entries from Excel")
        return schedules

    def _resolve_merged_cells(self, ws) -> List[List]:
        """
        Build a flat grid from a worksheet, forward-filling merged cell values
        so downstream code sees the value in every cell of a merged range.
        """
        merged_values: dict = {}
        for merged_range in ws.merged_cells.ranges:
            top_left_val = ws.cell(merged_range.min_row, merged_range.min_col).value
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_values[(r, c)] = top_left_val

        rows = []
        for r in range(1, ws.max_row + 1):
            row_data = []
            for c in range(1, ws.max_column + 1):
                val = merged_values.get((r, c), ws.cell(r, c).value)
                row_data.append(val)
            if any(v is not None for v in row_data):
                rows.append(row_data)
        return rows

    def _extract_from_rows(self, rows: List[List], source_name: str) -> List[ScheduleEntry]:
        """
        Extract schedule entries from a resolved list-of-rows.
        Handles any program code, not just BICS.
        """
        schedules = []
        current_class_group = None
        current_semester = None

        # Infer class group from sheet name as a fallback seed
        sheet_seed = self._infer_class_group_from_name(source_name)

        # Find the time-slot header row (first row with >1 time patterns)
        time_slots: List[Optional[str]] = []
        for row in rows[:15]:
            candidate = self._extract_time_slots_from_row(row)
            if sum(1 for t in candidate if t) > 1:
                time_slots = candidate
                break

        print(f"    ⏰ Detected time slots: {[t for t in time_slots if t]}")

        for row in rows:
            try:
                row_text = " ".join(str(v) for v in row if v is not None).strip()
                if not row_text:
                    continue

                # Try to detect a class-group header in this row
                class_info = self._detect_class_header_from_text(row_text)
                if class_info:
                    current_class_group, current_semester = class_info
                    print(f"    📋 Found class group: {current_class_group} ({current_semester})")
                    continue

                # Use sheet-name seed if no header found yet
                if current_class_group is None and sheet_seed:
                    current_class_group = sheet_seed

                # Try to detect a day row
                day = self._detect_day_from_first_cell(row)
                if day and current_class_group:
                    print(f"      📅 Processing {day}:")
                    for col_idx, cell_value in enumerate(row[1:], 1):
                        if cell_value is None or not str(cell_value).strip():
                            continue
                        time_slot = (
                            time_slots[col_idx - 1]
                            if col_idx - 1 < len(time_slots)
                            else None
                        )
                        if not time_slot:
                            continue
                        entry = self._parse_cell_content(
                            str(cell_value),
                            current_class_group,
                            day,
                            time_slot,
                            current_semester or "Unknown",
                        )
                        if entry:
                            schedules.append(entry)
                            print(f"        ⏰ {time_slot}: {entry.subject} in {entry.room}")

            except Exception as e:
                logger.warning(f"Error processing row in {source_name}: {e}")
                continue

        return schedules

    def _resolve_class_group_from_sheet(self, sheet_name: str) -> Optional[Tuple[str, str]]:
        """Try to extract class group and semester from a sheet name."""
        return self._detect_class_header_from_text(sheet_name)

    def _infer_class_group_from_name(self, name: str) -> Optional[str]:
        """Pull a bare class-group token (e.g. 'BICS 1A') from a sheet name."""
        match = re.search(r"([A-Z]{2,6}\s*\d[A-Z])", name, re.IGNORECASE)
        return match.group(1).upper() if match else None

    def _detect_class_header_from_text(self, text: str) -> Optional[Tuple[str, str]]:
        """
        Generalised class-group header detection.
        Supports any programme code (BICS, BCOM, BBIT, IT, MBA, BSC …).
        """
        sem = r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*(?:\s*[-–]\s*(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*)?\s+20\d{2}"

        # "BICS 1A Aug-Dec 2025" or "BCOM 2B Jan-May 2024"
        m = re.search(rf"([A-Z]{{2,6}}\s+\d[A-Z])\s+.*?({sem})", text, re.IGNORECASE)
        if m:
            return m.group(1).upper(), m.group(2)

        # "1A Aug-Dec 2025" (programme implied by sheet name)
        m = re.search(rf"\b(\d[A-Z])\b.*?({sem})", text, re.IGNORECASE)
        if m:
            return m.group(1).upper(), m.group(2)

        # "Year 1 Aug-Dec 2025"
        m = re.search(rf"((?:Year|Yr)\s+\d)\s+.*?({sem})", text, re.IGNORECASE)
        if m:
            return m.group(1), m.group(2)

        return None

    def _extract_time_slots_from_row(self, row: List) -> List[Optional[str]]:
        """Extract time slots from a raw row (list of cell values)."""
        result = []
        for val in row:
            if val is None:
                result.append(None)
                continue
            val_str = str(val).strip()
            found = False
            for pattern in self.time_patterns:
                m = re.search(pattern, val_str)
                if m:
                    sh, sm, eh, em = m.groups()
                    result.append(f"{sh.zfill(2)}:{sm}-{eh.zfill(2)}:{em}")
                    found = True
                    break
            if not found:
                result.append(None)
        return result

    def _detect_day_from_first_cell(self, row: List) -> Optional[str]:
        """Detect a weekday name in the first non-empty cell of a row."""
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        first = next((str(v) for v in row if v is not None), "")
        for day in days:
            if day.lower() in first.lower():
                return day
        return None

    def _extract_from_pdf(self, file_path: Path) -> List[ScheduleEntry]:
        """Extract schedule from PDF using table detection."""
        schedules = []

        try:
            # Extract all tables from PDF
            tables = tabula.read_pdf(
                file_path, pages="all", multiple_tables=True
            )

            for i, df in enumerate(tables):
                print(f"  📋 Processing PDF table {i+1}")
                table_schedules = self._extract_from_dataframe(
                    df, f"pdf_table_{i+1}"
                )
                schedules.extend(table_schedules)

        except Exception as e:
            logger.error(f"Error extracting tables from PDF: {e}")

        print(f"✅ Extracted {len(schedules)} schedule entries from PDF")
        return schedules

    def _extract_from_dataframe(
        self, df: pd.DataFrame, source_name: str
    ) -> List[ScheduleEntry]:
        """Extract schedule entries from DataFrame with context propagation."""
        schedules = []
        current_class_group = None
        current_semester = None

        print(
            f"    📊 Processing table with {len(df)} rows, {len(df.columns)} columns"
        )

        # Get time slots from column headers
        time_slots = self._extract_time_slots_from_columns(df.columns)
        print(f"    ⏰ Detected time slots: {time_slots}")

        for idx, row in df.iterrows():
            try:
                # Check for class group header
                class_info = self._detect_class_header(row)
                if class_info:
                    current_class_group, current_semester = class_info
                    print(
                        f"    📋 Found class group: {current_class_group} ({current_semester})"
                    )
                    continue

                # Check for day row
                day = self._detect_day_row(row)
                if day and current_class_group:
                    print(f"      📅 Processing {day}:")

                    # Process each time slot
                    for col_idx, cell_content in enumerate(
                        row[1:], 1
                    ):  # Skip first column (day)
                        if pd.notna(cell_content) and str(cell_content).strip():
                            time_slot = (
                                time_slots[col_idx - 1]
                                if col_idx - 1 < len(time_slots)
                                else None
                            )

                            if time_slot:
                                schedule_entry = self._parse_cell_content(
                                    str(cell_content),
                                    current_class_group,
                                    day,
                                    time_slot,
                                    current_semester or "Unknown",
                                )

                                if schedule_entry:
                                    schedules.append(schedule_entry)
                                    print(
                                        f"        ⏰ {time_slot}: {schedule_entry.subject} in {schedule_entry.room}"
                                    )

            except Exception as e:
                logger.warning(f"Error processing row {idx}: {e}")
                continue

        return schedules

    def _extract_time_slots_from_columns(self, columns) -> List[str]:
        """Extract time slots from DataFrame column headers."""
        time_slots = []

        for col in columns:
            if pd.isna(col):
                time_slots.append(None)
                continue

            col_str = str(col).strip()

            # Look for time patterns
            for pattern in self.time_patterns:
                match = re.search(pattern, col_str)
                if match:
                    start_hour, start_min, end_hour, end_min = match.groups()
                    time_slot = f"{start_hour.zfill(2)}:{start_min}-{end_hour.zfill(2)}:{end_min}"
                    time_slots.append(time_slot)
                    break
            else:
                time_slots.append(None)

        return time_slots

    def _detect_class_header(self, row) -> Optional[Tuple[str, str]]:
        """Detect if a DataFrame row contains a class group header."""
        row_text = " ".join(str(cell) for cell in row if pd.notna(cell)).strip()
        return self._detect_class_header_from_text(row_text)

    def _detect_day_row(self, row) -> Optional[str]:
        """Detect if row represents a day of the week."""
        first_cell = (
            str(row.iloc[0]) if len(row) > 0 and pd.notna(row.iloc[0]) else ""
        )

        days = [
            "Monday",
            "Tuesday",
            "Wednesday",
            "Thursday",
            "Friday",
            "Saturday",
            "Sunday",
        ]

        for day in days:
            if day.lower() in first_cell.lower():
                return day

        return None

    def _parse_cell_content(
        self,
        content: str,
        class_group: str,
        day: str,
        time_slot: str,
        semester: str,
    ) -> Optional[ScheduleEntry]:
        """Parse individual cell content using LLM for intelligent extraction."""

        if not content.strip() or content.strip().upper() in [
            "NAN",
            "NONE",
            "",
        ]:
            return None

        # Try pattern-based extraction first (faster)
        pattern_result = self._try_pattern_extraction(
            content, class_group, day, time_slot, semester
        )
        if pattern_result:
            return pattern_result

        # Fall back to LLM extraction for complex cases
        return self._llm_parse_cell_content(
            content, class_group, day, time_slot, semester
        )

    def _try_pattern_extraction(
        self,
        content: str,
        class_group: str,
        day: str,
        time_slot: str,
        semester: str,
    ) -> Optional[ScheduleEntry]:
        """Try to extract using regex patterns before using LLM."""

        # Extract unit code — any 2-5 uppercase letters followed by 3-4 digits
        unit_code_match = re.search(r"\b([A-Z]{2,5}\s*\d{3,4})\b", content)
        unit_code = unit_code_match.group(1).replace(" ", "") if unit_code_match else None

        # Extract room
        room = None
        for pattern in self.room_patterns:
            match = re.search(pattern, content, re.IGNORECASE)
            if match:
                room = match.group(0)
                break

        # Extract instructor (Dr. Name pattern)
        instructor_match = re.search(
            r"Dr\.?\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)", content, re.IGNORECASE
        )
        instructor = instructor_match.group(0) if instructor_match else None

        # If we got the key components, create entry
        if unit_code and room:
            # Extract subject (everything before unit code or room)
            subject_parts = re.split(
                r"(?:"
                + "|".join(self.room_patterns)
                + r")|(?:ICS|HED|MATH|STAT)\s*\d{4}",
                content,
                flags=re.IGNORECASE,
            )
            subject = (
                subject_parts[0].strip() if subject_parts else "Unknown Subject"
            )

            start_time, end_time = self._parse_time_slot(time_slot)

            return ScheduleEntry(
                class_group=class_group,
                subject=subject,
                unit_code=unit_code,
                room=room,
                day=day,
                start_time=start_time,
                end_time=end_time,
                instructor=instructor or "Unknown Instructor",
                session_type=self._determine_session_type(content),
                semester=semester,
                raw_content=content,
                confidence=0.8,
            )

        return None

    def _llm_parse_cell_content(
        self,
        content: str,
        class_group: str,
        day: str,
        time_slot: str,
        semester: str,
    ) -> Optional[ScheduleEntry]:
        """Use LLM to parse complex cell content."""

        try:
            prompt = f"""
            Extract schedule information from this cell content and return as JSON:
            
            Content: "{content}"
            Context: Class {class_group}, {day} {time_slot}, {semester}
            
            Extract and return valid JSON only (no other text):
            {{
                "subject": "subject name or null",
                "room": "room/location or null", 
                "instructor": "instructor name or null",
                "unit_code": "course code like ICS1202 or null",
                "session_type": "lecture/lab/practical or null"
            }}
            
            Rules:
            - If any field cannot be determined, use null
            - Subject should be the main course name
            - Room should be location/room identifier
            - Instructor should be the teacher's name
            - Unit code should be the course code (like ICS1202)
            - Session type should be lecture, lab, or practical based on context
            """

            response = self.openai_client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[{"role": "user", "content": prompt}],
                temperature=0.1,
                max_tokens=200,
            )

            result_text = response.choices[0].message.content.strip()

            # Extract JSON from response (in case there's extra text)
            json_match = re.search(r"\{.*\}", result_text, re.DOTALL)
            if json_match:
                result_text = json_match.group(0)

            parsed = json.loads(result_text)

            # Validate required fields
            if not parsed.get("subject") and not parsed.get("unit_code"):
                return None

            start_time, end_time = self._parse_time_slot(time_slot)

            return ScheduleEntry(
                class_group=class_group,
                subject=parsed.get("subject", "Unknown Subject"),
                unit_code=parsed.get("unit_code"),
                room=parsed.get("room", "Unknown Room"),
                day=day,
                start_time=start_time,
                end_time=end_time,
                instructor=parsed.get("instructor", "Unknown Instructor"),
                session_type=parsed.get("session_type", "lecture"),
                semester=semester,
                raw_content=content,
                confidence=0.9,
            )

        except Exception as e:
            logger.warning(f"LLM parsing failed for content '{content}': {e}")
            return None

    def _parse_time_slot(self, time_slot: str) -> Tuple[str, str]:
        """Parse time slot into start and end times."""
        if not time_slot:
            return "00:00", "00:00"

        for pattern in self.time_patterns:
            match = re.search(pattern, time_slot)
            if match:
                start_hour, start_min, end_hour, end_min = match.groups()
                start_time = f"{start_hour.zfill(2)}:{start_min}"
                end_time = f"{end_hour.zfill(2)}:{end_min}"
                return start_time, end_time

        return "00:00", "00:00"

    def _determine_session_type(self, content: str) -> str:
        """Determine session type from content."""
        content_lower = content.lower()

        if any(
            lab_word in content_lower
            for lab_word in ["lab", "laboratory", "practical"]
        ):
            return "lab"
        elif any(lec_word in content_lower for lec_word in ["lecture", "lt"]):
            return "lecture"
        else:
            return "lecture"  # Default

    def convert_to_semantic_chunks(
        self, schedules: List[ScheduleEntry]
    ) -> List[Dict[str, Any]]:
        """Convert schedule entries to semantic chunks for hybrid search."""
        chunks = []

        for i, entry in enumerate(schedules):
            # Create natural language description
            chunk_text = f"{entry.class_group} has {entry.subject}"
            if entry.unit_code:
                chunk_text += f" ({entry.unit_code})"
            chunk_text += f" in {entry.room} on {entry.day} from {entry.start_time} to {entry.end_time}"
            if entry.instructor:
                chunk_text += f" with {entry.instructor}"
            chunk_text += f". This is a {entry.session_type} session during {entry.semester}."

            chunk_id = hashlib.md5(
                f"{entry.class_group}_{entry.day}_{entry.start_time}_{entry.subject}".encode()
            ).hexdigest()

            chunk = {
                "chunk_id": chunk_id,
                "text": chunk_text,
                "metadata": {
                    "doc_type": "schedule",
                    "is_schedule": True,
                    "class_group": entry.class_group,
                    "subject": entry.subject,
                    "day": entry.day,
                    "time_slot": f"{entry.start_time}-{entry.end_time}",
                    "room": entry.room,
                    "instructor": entry.instructor,
                    "semester": entry.semester,
                    "session_type": entry.session_type,
                    "unit_code": entry.unit_code,
                    "extraction_confidence": entry.confidence,
                },
            }
            chunks.append(chunk)

        return chunks
